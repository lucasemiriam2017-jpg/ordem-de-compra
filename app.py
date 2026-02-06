from flask import (
    Flask, render_template, request, send_file, redirect, url_for,
    session, Response
)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from psycopg.rows import dict_row
from contextlib import contextmanager
from datetime import datetime, date
from dotenv import load_dotenv

import psycopg
from openpyxl import load_workbook
import io, os, re, urllib.parse, json, csv

# -------------------- CONFIG --------------------
load_dotenv()
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret")

DATABASE_URL = os.environ.get("DATABASE_URL")
ADMIN_USER = os.environ.get("ADMIN_USER")
ADMIN_PASS = os.environ.get("ADMIN_PASS")

LOGO_PATH = os.path.join("static", "logo.png")
PDF_PREFIX = "Ordem_Compra"


# -------------------- DB --------------------
@contextmanager
def get_conn_cursor():
    # row_factory=dict_row => fetchone()/fetchall() retornam dict
    conn = psycopg.connect(DATABASE_URL, row_factory=dict_row)
    try:
        cur = conn.cursor()
        yield conn, cur
        conn.commit()
    finally:
        conn.close()


def only_digits(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def parse_date_safe(v):
    """
    Converte para date se possível:
    - datetime/date já ok
    - string tenta parse ISO
    - outros retorna None
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # tenta ISO yyyy-mm-dd
        try:
            return datetime.fromisoformat(s).date()
        except Exception:
            pass
        # tenta dd/mm/yyyy
        try:
            return datetime.strptime(s, "%d/%m/%Y").date()
        except Exception:
            return None
    return None


# -------------------- XLSX HELPERS --------------------
def xlsx_iter_rows(file_storage):
    """
    Lê primeira aba do XLSX e retorna (header:list[str], data:list[tuple])
    """
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = rows[1:]
    return header, data


def get_col_index(header, name):
    name_l = name.strip().lower()
    for i, h in enumerate(header):
        if (h or "").strip().lower() == name_l:
            return i
    return None


def to_date(v):
    # openpyxl normalmente já entrega date/datetime
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # tenta converter string
    return parse_date_safe(v)


# -------------------- TABELAS --------------------
with get_conn_cursor() as (conn, cursor):
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS empresas_cnpj (
        cnpj TEXT PRIMARY KEY,
        bp TEXT,
        raw_json JSONB,
        updated_at TIMESTAMP DEFAULT NOW()
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS partidas_aberto (
        id INTEGER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
        bp TEXT,
        data_base DATE,
        compensacao TEXT,
        valor NUMERIC,
        raw_json JSONB,
        updated_at TIMESTAMP DEFAULT NOW()
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS ordens_compra (
        id INTEGER GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
        created_at TIMESTAMP DEFAULT NOW(),

        cliente_json JSONB,
        filial_json JSONB,
        itens_json JSONB,

        cliente_nome TEXT,
        cliente_cnpj TEXT,
        bp TEXT,

        status_auto TEXT,
        status_manual TEXT,
        status_final TEXT,
        status_detail JSONB,

        pdf_nome TEXT,
        pdf_bytes BYTEA
    );
    """)


# -------------------- REGRAS --------------------
def buscar_bp_por_cnpj(cnpj_digits: str):
    if not cnpj_digits:
        return None
    with get_conn_cursor() as (conn, cursor):
        cursor.execute("SELECT bp FROM empresas_cnpj WHERE cnpj = %s", (cnpj_digits,))
        row = cursor.fetchone()
        return row["bp"] if row else None


def buscar_partidas_em_aberto(bp: str):
    # em aberto = Compensaç vazia
    with get_conn_cursor() as (conn, cursor):
        cursor.execute("""
            SELECT * FROM partidas_aberto
            WHERE bp = %s
              AND (compensacao IS NULL OR TRIM(compensacao) = '')
        """, (bp,))
        return cursor.fetchall()


def calcular_status(cnpj_digits: str):
    bp = buscar_bp_por_cnpj(cnpj_digits)
    hoje = date.today()

    detail = {"cnpj": cnpj_digits, "bp": bp, "motivo": None,
              "resumo": {"abertas": 0, "vencidas": 0, "a_vencer": 0}}

    if not bp:
        detail["motivo"] = "CNPJ não encontrado no Relatório 1 (empresas_cnpj)."
        return "NÃO POSSUI CADASTRO", None, detail

    partidas = buscar_partidas_em_aberto(bp)
    vencidas = 0
    a_vencer = 0

    for p in partidas:
        db = p.get("data_base")
        if db and db < hoje:
            vencidas += 1
        else:
            a_vencer += 1

    detail["resumo"]["abertas"] = len(partidas)
    detail["resumo"]["vencidas"] = vencidas
    detail["resumo"]["a_vencer"] = a_vencer

    if vencidas > 0:
        detail["motivo"] = "Existe(m) partida(s) em aberto vencida(s) (Data base < hoje)."
        return "NÃO LIBERAR", bp, detail

    if a_vencer > 0:
        detail["motivo"] = "Existe(m) partida(s) em aberto a vencer (Data base >= hoje)."
        return "VERIFICAR COM FINANCEIRO", bp, detail

    detail["motivo"] = "Sem partidas em aberto."
    return "LIBERAR", bp, detail


# -------------------- ROTAS PÚBLICAS --------------------
@app.route("/")
def index():
    return render_template("index.html")  # seu layout atual


@app.route("/gerar_pdf", methods=["POST"])
def gerar_pdf():
    data = request.json or {}
    cliente = data.get("cliente", {}) or {}
    filial = data.get("filial", {}) or {}
    itens = data.get("itens", []) or []
    obs = data.get("obs", "")
    pagamento = data.get("pagamento", "")
    prazo = data.get("prazo", "")

    if not itens:
        return "Erro: Nenhum item adicionado.", 400

    cnpj_digits = only_digits(cliente.get("CNPJ", ""))
    empresa_nome = cliente.get("Empresa", "Documento")

    status_auto, bp, detail = calcular_status(cnpj_digits)

    # ----- gera PDF (mesmo padrão do seu app atual) -----
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.2 * cm, bottomMargin=2 * cm
    )

    s = getSampleStyleSheet()
    st = {
        "title": ParagraphStyle("title", parent=s["Heading1"], alignment=TA_CENTER, fontSize=16),
        "n": ParagraphStyle("n", parent=s["Normal"], fontSize=9),
        "small": ParagraphStyle("small", parent=s["Normal"], fontSize=8),
        "tabela": ParagraphStyle("tabela", parent=s["Normal"], fontSize=8.5, alignment=TA_LEFT),
        "center": ParagraphStyle("center", parent=s["Normal"], fontSize=10, alignment=TA_CENTER, spaceAfter=6)
    }

    e = []
    if os.path.exists(LOGO_PATH):
        logo = Image(LOGO_PATH, width=7.5 * cm, height=2.3 * cm)
        logo.hAlign = "CENTER"
        e += [logo, Spacer(1, 4)]

    e += [Paragraph("<b>ORDEM DE COMPRA</b>", st["title"]), Spacer(1, 6)]

    e.append(Paragraph("<b>EMPRESA SOLICITANTE</b>", st["center"]))
    cliente_data = [[Paragraph(f"<b>{k}</b>", st["n"]), Paragraph(str(v), st["n"])] for k, v in cliente.items()]
    tabela_cliente = Table(cliente_data, colWidths=[4 * cm, 11 * cm])
    tabela_cliente.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.4, colors.grey)]))
    e += [tabela_cliente, Spacer(1, 10)]

    e.append(Paragraph("<b>FILIAL / FORNECEDOR</b>", st["center"]))
    filial_data = [[Paragraph(f"<b>{k}</b>", st["n"]), Paragraph(str(v), st["n"])] for k, v in filial.items()]
    tabela_filial = Table(filial_data, colWidths=[4 * cm, 11 * cm])
    tabela_filial.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.4, colors.grey)]))
    e += [tabela_filial, Spacer(1, 12)]

    e.append(Paragraph(f"<b>Prazo de Entrega:</b> {prazo}", st["n"]))
    e.append(Spacer(1, 16))

    e.append(Paragraph("<b>LISTAGEM DE PRODUTOS</b>", st["center"]))
    e.append(Spacer(1, 8))

    cols = [1.2*cm, 1.0*cm, 3.2*cm, 8.3*cm, 3.1*cm, 2.3*cm]
    data_table = [["ITEM", "QTD", "CÓDIGO", "DESCRIÇÃO", "PREÇO UNIT (R$)", "TOTAL (R$)"]]
    total_geral = 0

    for i, item in enumerate(itens, start=1):
        qtd = float(item.get("qtd", 0) or 0)
        cod = item.get("cod", "")
        desc = item.get("desc", "")
        preco = float(str(item.get("preco", "0")).replace(".", "").replace(",", ".") or 0)
        tot = float(str(item.get("tot", "0")).replace(".", "").replace(",", ".") or 0)
        total_geral += tot
        p_fmt = f"R$ {preco:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        t_fmt = f"R$ {tot:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        data_table.append([str(i), qtd, cod, Paragraph(desc, st["tabela"]), p_fmt, t_fmt])

    total_fmt = f"{total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    data_table.append(["", "", "", Paragraph("<b>TOTAL GERAL</b>", st["tabela"]), "", f"R$ {total_fmt}"])

    t = Table(data_table, colWidths=cols, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#004C99")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (3, 1), (3, -1), "LEFT"),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    e += [t, Spacer(1, 20)]

    e.append(Paragraph(f"<b>Condição de Pagamento:</b> Boleto em {pagamento}", st["n"]))
    e.append(Spacer(1, 10))

    if obs:
        e.append(Paragraph("<b>OBSERVAÇÕES:</b>", st["n"]))
        e.append(Paragraph(obs, st["n"]))
        e.append(Spacer(1, 12))

    e.append(Paragraph("A ORDEM DE COMPRA DEVE SER ENVIADA PARA <b>convenios@farmaciassaojoao.com.br</b>", st["small"]))
    e.append(Paragraph("<i>*A via original deve ser entregue na filial da venda*</i>", st["small"]))
    e.append(Spacer(1, 36))
    e.append(Paragraph("Assinatura e carimbo: _________________________________", st["n"]))

    doc.build(e)
    buffer.seek(0)
    pdf_bytes = buffer.getvalue()

    nome_arquivo = f"{PDF_PREFIX}_{str(empresa_nome).replace(' ', '_')}.pdf"

    with get_conn_cursor() as (conn, cursor):
        cursor.execute("""
            INSERT INTO ordens_compra (
                cliente_json, filial_json, itens_json,
                cliente_nome, cliente_cnpj, bp,
                status_auto, status_manual, status_final, status_detail,
                pdf_nome, pdf_bytes
            ) VALUES (%s,%s,%s, %s,%s,%s, %s,NULL,%s,%s, %s,%s)
        """, (
            json.dumps(cliente, ensure_ascii=False),
            json.dumps(filial, ensure_ascii=False),
            json.dumps(itens, ensure_ascii=False),
            empresa_nome, cnpj_digits, bp,
            status_auto, status_auto,
            json.dumps(detail, ensure_ascii=False),
            nome_arquivo, pdf_bytes  # psycopg v3 aceita bytes direto
        ))

    return send_file(
        io.BytesIO(pdf_bytes),
        as_attachment=True,
        download_name=urllib.parse.quote(nome_arquivo),
        mimetype="application/pdf"
    )


# -------------------- ADMIN --------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario")
        senha = request.form.get("senha")
        if usuario == ADMIN_USER and senha == ADMIN_PASS:
            session["logged_in"] = True
            return redirect(url_for("lista_ordens"))
        return render_template("login.html", error="❌ Usuário ou senha inválidos")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect(url_for("login"))


def require_login():
    return session.get("logged_in")


@app.route("/admin/ordens")
def lista_ordens():
    if not require_login():
        return redirect(url_for("login"))
    with get_conn_cursor() as (conn, cursor):
        cursor.execute("""
            SELECT id, created_at, cliente_nome, cliente_cnpj, bp,
                   status_auto, status_manual, status_final
            FROM ordens_compra
            ORDER BY id DESC
        """)
        ordens = cursor.fetchall()
    return render_template("ordens.html", ordens=ordens)


@app.route("/admin/ordens/<int:ordem_id>/pdf")
def ordem_pdf(ordem_id):
    if not require_login():
        return redirect(url_for("login"))
    with get_conn_cursor() as (conn, cursor):
        cursor.execute("SELECT pdf_nome, pdf_bytes FROM ordens_compra WHERE id=%s", (ordem_id,))
        row = cursor.fetchone()
    if not row or not row.get("pdf_bytes"):
        return "PDF não encontrado", 404
    return send_file(
        io.BytesIO(row["pdf_bytes"]),
        download_name=row["pdf_nome"],
        as_attachment=True,
        mimetype="application/pdf"
    )


@app.route("/admin/ordens/<int:ordem_id>/status", methods=["POST"])
def set_status(ordem_id):
    if not require_login():
        return redirect(url_for("login"))
    novo = (request.form.get("status_manual") or "").strip() or None
    with get_conn_cursor() as (conn, cursor):
        if novo is None:
            cursor.execute("""
                UPDATE ordens_compra
                SET status_manual=NULL,
                    status_final=status_auto
                WHERE id=%s
            """, (ordem_id,))
        else:
            cursor.execute("""
                UPDATE ordens_compra
                SET status_manual=%s,
                    status_final=%s
                WHERE id=%s
            """, (novo, novo, ordem_id))
    return redirect(url_for("lista_ordens"))


@app.route("/admin/ordens/baixar-csv")
def baixar_csv():
    if not require_login():
        return redirect(url_for("login"))
    with get_conn_cursor() as (conn, cursor):
        cursor.execute("""
            SELECT id, created_at, cliente_nome, cliente_cnpj, bp,
                   status_auto, status_manual, status_final
            FROM ordens_compra
            ORDER BY id DESC
        """)
        rows = cursor.fetchall()

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["ID", "Data", "Empresa", "CNPJ", "BP", "Status Auto", "Status Manual", "Status Final"])
    for r in rows:
        w.writerow([r["id"], r["created_at"], r["cliente_nome"], r["cliente_cnpj"], r["bp"],
                    r["status_auto"], r["status_manual"], r["status_final"]])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=ordens_compra.csv"}
    )


@app.route("/admin/importar-cnpj", methods=["POST"])
def importar_cnpj():
    if not require_login():
        return redirect(url_for("login"))

    arq = request.files.get("arquivo")
    if not arq or arq.filename == "":
        return "Arquivo não enviado", 400

    header, data = xlsx_iter_rows(arq)

    idx_cnpj = get_col_index(header, "Número CNPJ")
    idx_bp = get_col_index(header, "Cliente")

    if idx_cnpj is None or idx_bp is None:
        return f"Preciso das colunas 'Número CNPJ' e 'Cliente'. Colunas: {header}", 400

    with get_conn_cursor() as (conn, cursor):
        for r in data:
            cnpj = only_digits(str(r[idx_cnpj] or ""))
            bp = str(r[idx_bp] or "").strip()
            if not cnpj or not bp:
                continue

            raw = {header[i]: r[i] for i in range(len(header))}
            cursor.execute("""
                INSERT INTO empresas_cnpj (cnpj, bp, raw_json, updated_at)
                VALUES (%s,%s,%s,NOW())
                ON CONFLICT (cnpj) DO UPDATE SET
                    bp=EXCLUDED.bp,
                    raw_json=EXCLUDED.raw_json,
                    updated_at=NOW()
            """, (cnpj, bp, json.dumps(raw, ensure_ascii=False, default=str)))

    return redirect(url_for("lista_ordens"))


@app.route("/admin/importar-partidas", methods=["POST"])
def importar_partidas():
    if not require_login():
        return redirect(url_for("login"))

    arq = request.files.get("arquivo")
    if not arq or arq.filename == "":
        return "Arquivo não enviado", 400

    header, data = xlsx_iter_rows(arq)

    idx_bp = get_col_index(header, "Cliente")
    idx_db = get_col_index(header, "Data base")
    idx_comp = get_col_index(header, "Compensaç.")
    idx_val = get_col_index(header, "Montante em MI")  # opcional

    if idx_bp is None or idx_db is None or idx_comp is None:
        return f"Preciso das colunas 'Cliente', 'Data base', 'Compensaç.'. Colunas: {header}", 400

    with get_conn_cursor() as (conn, cursor):
        cursor.execute("DELETE FROM partidas_aberto")

        for r in data:
            bp = str(r[idx_bp] or "").strip()
            if not bp:
                continue

            data_base = to_date(r[idx_db])
            comp = str(r[idx_comp] or "").strip()
            valor = None
            if idx_val is not None:
                try:
                    valor = float(r[idx_val]) if r[idx_val] is not None else None
                except Exception:
                    valor = None

            raw = {header[i]: r[i] for i in range(len(header))}
            cursor.execute("""
                INSERT INTO partidas_aberto (bp, data_base, compensacao, valor, raw_json, updated_at)
                VALUES (%s,%s,%s,%s,%s,NOW())
            """, (bp, data_base, comp, valor, json.dumps(raw, ensure_ascii=False, default=str)))

    return redirect(url_for("lista_ordens"))


if __name__ == "__main__":
    app.run(debug=True)
